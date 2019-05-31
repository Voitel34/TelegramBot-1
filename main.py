# _*_ coding:utf-8 _*_

import time     # подключение библиотек для работы со временем
import datetime
import telebot  # подключение библиотеки для работы с API Telegram
import constants  # подключение новосозданной библиотеки констант
from openpyxl import load_workbook  # подключение библиотеки для работы с электронными таблицами

bot = telebot.TeleBot(constants.token)  # создание бота
log_file = load_workbook(filename='log.xlsx', data_only=True)  # привязка электронных таблиц
users_id_file = load_workbook(filename='users_id.xlsx', data_only=True)
sheet = log_file['Лист1']  # привязка листов таблиц
sheet_id = users_id_file['Лист1']


@bot.message_handler(commands=['start'])  # обработчик команды (/start)
def handle_text(message):
    bot.send_message(message.chat.id, constants.text_for_start)


@bot.message_handler(commands=['supermessage'])     # обработчик команды (/supermessage)
def handle_text(message):
    if message.chat.id == '9999999':    # проверка id администратора
        row_number = 2      # стартовая строка таблицы
        while sheet_id['B' + str(row_number)].value is not None:    # проход по всем существующим id
            bot.send_message(sheet_id['B' + str(row_number)].value, constants.text_for_notifications)
            row_number += 1
            time.sleep(5)


@bot.message_handler(commands=['registration'])     # обработчик команды (/registration)
def handle_text(message):
    sent = bot.send_message(message.chat.id, 'Введите фамилию и имя в формате: Иванов Иван')
    bot.register_next_step_handler(sent, user_name)  # переход к следующей функции


@bot.message_handler(commands=['menu'])     # обработчик команды (/menu)
def handle_text(message):
    user_markup = telebot.types.ReplyKeyboardMarkup(True, False)
    user_markup.row('Я оплатил', 'Я задерживаю оплату')
    user_markup.row('Сколько я должен?', 'Закрыть клавиатуру')
    bot.send_message(message.chat.id, 'Меню открыто', reply_markup=user_markup)


@bot.message_handler(content_types=['text'])    # обработчик текстового сообщения
def handle_text(message):
    if message.text == 'Закрыть клавиатуру':
        hide_markup = telebot.types.ReplyKeyboardRemove()
        bot.send_message(message.chat.id, 'Меню закрыто', reply_markup=hide_markup)
    if message.text == 'Я оплатил':
        user_markup = telebot.types.ReplyKeyboardMarkup(True, False)
        user_markup.row('2019')
        user_markup.row('Закрыть клавиатуру')
        sent = bot.send_message(message.chat.id, 'Выберите год', reply_markup=user_markup)
        bot.register_next_step_handler(sent, choose_year_for_point)
    if message.text == 'Я задерживаю оплату':
        user_markup = telebot.types.ReplyKeyboardMarkup(True, False)
        user_markup.row('2019')
        user_markup.row('Закрыть клавиатуру')
        sent = bot.send_message(message.chat.id, 'Выберите год', reply_markup=user_markup)
        bot.register_next_step_handler(sent, choose_year_for_delay_point)
    if message.text == 'Сколько я должен?':
        debt = 0
        row_number_id = 2
        while sheet_id['B' + str(row_number_id)].value != message.chat.id:
            row_number_id += 1
        name_of_user = sheet_id['A' + str(row_number_id)].value
        row_number = 3
        while sheet['A' + str(row_number)].value != name_of_user:
            row_number += 1
        if datetime.datetime.now().day < 23:
            for i in range(6, datetime.datetime.now().month):
                if str(sheet.cell(row=row_number, column=i.value)) == '' or str(sheet.cell(row=row_number, column=i.value)) == 'отсрочка':
                    debt += 100
        else:
            for i in range(6, datetime.datetime.now().month + 1):
                if str(sheet.cell(row=row_number, column=i).value) == '' or str(sheet.cell(row=row_number, column=i).value) == 'отсрочка':
                    debt += 100
        bot.send_message(message.chat.id, constants.text_for_debt.format(debt))


def user_name(message):
    row_number = 3
    count = 0
    while sheet['A' + str(row_number)].value is not None:
        if sheet['A' + str(row_number)].value == message.text.strip().title():
            count += 1
        row_number += 1
    if count:
        bot.send_message(message.chat.id, 'Вы уже зарегистрированы')
    else:
        sheet['A' + str(row_number)].value = message.text.strip().title()
        sheet_id['A' + str(row_number - 1)].value = message.text.strip().title()
        sheet_id['B' + str(row_number - 1)].value = message.chat.id
        bot.send_message(message.chat.id, 'Регистрация прошла успешно')
    log_file.save('log.xlsx')
    users_id_file.save('users_id.xlsx')


def choose_year_for_point(message):
    if message.text == '2019':
        user_markup = telebot.types.ReplyKeyboardMarkup(True, False)
        user_markup.row('июнь', 'июль', 'август')
        user_markup.row('сентябрь', 'октябрь', 'ноябрь', 'декабрь')
        user_markup.row('Закрыть клавиатуру')
        sent = bot.send_message(message.chat.id, 'Выберите месяц', reply_markup=user_markup)
        bot.register_next_step_handler(sent, choose_month_for_point)


def choose_month_for_point(message):
    if message.text == 'январь':
        point(1, message.chat.id)
    elif message.text == 'февраль':
        point(2, message.chat.id)
    elif message.text == 'март':
        point(3, message.chat.id)
    elif message.text == 'апрель':
        point(4, message.chat.id)
    elif message.text == 'май':
        point(5, message.chat.id)
    elif message.text == 'июнь':
        point(6, message.chat.id)
    elif message.text == 'июль':
        point(7, message.chat.id)
    elif message.text == 'август':
        point(8, message.chat.id)
    elif message.text == 'сентябрь':
        point(9, message.chat.id)
    elif message.text == 'октябрь':
        point(10, message.chat.id)
    elif message.text == 'ноябрь':
        point(11, message.chat.id)
    elif message.text == 'декабрь':
        point(12, message.chat.id)


def point(month, id_of_user):
    row_number_id = 2
    while sheet_id['B' + str(row_number_id)].value != id_of_user:
        row_number_id += 1
    name_of_user = sheet_id['A' + str(row_number_id)].value
    row_number = 3
    while sheet['A' + str(row_number)].value != name_of_user:
        row_number += 1
    sheet.cell(row=row_number, column=(1 + month)).value = '+'
    bot.send_message(id_of_user, 'Отмечено')
    log_file.save('log.xlsx')


def choose_year_for_delay_point(message):
    if message.text == '2019':
        user_markup = telebot.types.ReplyKeyboardMarkup(True, False)
        user_markup.row('июнь', 'июль', 'август')
        user_markup.row('сентябрь', 'октябрь', 'ноябрь', 'декабрь')
        user_markup.row('Закрыть клавиатуру')
        sent = bot.send_message(message.chat.id, 'Выберите месяц', reply_markup=user_markup)
        bot.register_next_step_handler(sent, choose_month_for_delay_point)


def choose_month_for_delay_point(message):
    if message.text == 'январь':
        delay_point(1, message.chat.id)
    elif message.text == 'февраль':
        delay_point(2, message.chat.id)
    elif message.text == 'март':
        delay_point(3, message.chat.id)
    elif message.text == 'апрель':
        delay_point(4, message.chat.id)
    elif message.text == 'май':
        delay_point(5, message.chat.id)
    elif message.text == 'июнь':
        delay_point(6, message.chat.id)
    elif message.text == 'июль':
        delay_point(7, message.chat.id)
    elif message.text == 'август':
        delay_point(8, message.chat.id)
    elif message.text == 'сентябрь':
        delay_point(9, message.chat.id)
    elif message.text == 'октябрь':
        delay_point(10, message.chat.id)
    elif message.text == 'ноябрь':
        delay_point(11, message.chat.id)
    elif message.text == 'декабрь':
        delay_point(12, message.chat.id)


def delay_point(month, id_of_user):
    row_number_id = 2
    while sheet_id['B' + str(row_number_id)].value != id_of_user:
        row_number_id += 1
    name_of_user = sheet_id['A' + str(row_number_id)].value
    row_number = 3
    while sheet['A' + str(row_number)].value != name_of_user:
        row_number += 1
    sheet.cell(row=row_number, column=1 + month).value = 'отсрочка'
    bot.send_message(id_of_user, 'Отмечено')
    log_file.save('log.xlsx')


while True:
    try:
        bot.polling(none_stop=True)
    except Exception as e:
        print(e)
        time.sleep(15)
